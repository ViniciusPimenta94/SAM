import openpyxl
# from datetime import datetime

class consolidando_arquivos:
    def __init__(self, wb_fornecedor, Cliente, palavra_chave):
        print('Executando SAM_3...')
        self.Cliente = Cliente
        self.palavra_chave = palavra_chave
                
        self.wb_fornecedor = wb_fornecedor
        self.ws_fornecedor = self.wb_fornecedor['Worksheet2']
        
        self.wb_twm = openpyxl.load_workbook('./arquivo_twm/arquivo_TWM_'+self.Cliente+'.xlsx')
        try:
            self.ws_twm = self.wb_twm['Planilha1']
        except:
            self.ws_twm = self.wb_twm['faturas']
        
        self.wb_categoria = openpyxl.load_workbook('./arquivo_cliente/Categorias_Subcategoria.xlsx')
        
        self.ws_categoria = self.wb_categoria['Categorias']
        
        self.wb_localidade_riachuelo = openpyxl.load_workbook('./arquivo_cliente/Localidades_RIACHUELO.xlsx')
        self.ws_localidade_riachuelo = self.wb_localidade_riachuelo['Localidades']
                
    def ler_arquivo_consolidado(self):
        self.arquivo_consolidado = './arquivo_cliente/arquivo_consolidado_todos.xlsx'
   
    def ler_arquivo_linkado(self):
        self.arquivo_linkado = './arquivo_cliente/arquivo_linkado_todos.xlsx'
 
    def load_wb_Consolidado(self):
        wb_consolidado = openpyxl.load_workbook(self.arquivo_consolidado)
        nome_ws_consolidado_cliente = self.Cliente.lower()
        
        ws_consolidado = wb_consolidado[f'{nome_ws_consolidado_cliente}']
        ws_consolidado.title = 'Worksheet'
       
        
        self.wb_consolidado = wb_consolidado
        self.ws_consolidado = ws_consolidado

    def load_wb_linkado(self):
        wb_linkado = openpyxl.load_workbook(self.arquivo_linkado)
        nome_ws_linkado_cliente = self.Cliente.lower()
        
        try:
            ws_linkado = wb_linkado[f'{nome_ws_linkado_cliente}']
        except:
            print('Planilha do cliente linkado não encontrada!!!')
        
        self.wb_linkado = wb_linkado
        self.ws_linkado = ws_linkado

    def count_colunas(self):
        is_data = True
        count_col_consolidado = 1
        while is_data:
            count_col_consolidado += 1
            data =  self.ws_consolidado.cell (row = 1, column = count_col_consolidado).value
            if data == None:
                is_data = False
        count_col_consolidado-=1
        self.count_col_consolidado = count_col_consolidado
        
        is_data = True
        count_col_fornecedor = 1
        while is_data:
            count_col_fornecedor += 1
            data =  self.ws_fornecedor.cell (row = 1, column = count_col_fornecedor).value
            if data == None:
                is_data = False
        count_col_fornecedor-=1
        self.count_col_fornecedor = count_col_fornecedor
        
        is_data = True
        count_col_twm = 1
        while is_data:
            count_col_twm += 1
            data =  self.ws_twm.cell (row = 1, column = count_col_twm).value
            if data == None:
                is_data = False
        count_col_twm-=1
        self.count_col_twm = count_col_twm                

    def count_linhas(self):
        is_data = True
        count_row_fornecedor = 1
        while is_data:
            count_row_fornecedor += 1
            data =  self.ws_fornecedor.cell (row = count_row_fornecedor, column = 1).value
            if data == None:
                is_data = False
        count_row_fornecedor -=1
        self.count_row_fornecedor = count_row_fornecedor
        
        is_data = True
        count_row_linkado = 1
        while is_data:
            count_row_linkado += 1
            data =  self.ws_linkado.cell (row = count_row_linkado, column = 1).value
            if data == None:
                is_data = False
        count_row_linkado -=1
        self.count_row_linkado = count_row_linkado
        
        is_data = True
        count_row_twm = 1
        while is_data:
            count_row_twm += 1
            data =  self.ws_twm.cell (row = count_row_twm, column = 1).value
            if data == None:
                is_data = False
        count_row_twm -=1
        self.count_row_twm = count_row_twm
        
        is_data = True
        count_row_categoria = 1
        while is_data:
            count_row_categoria += 1
            data =  self.ws_categoria.cell (row = count_row_categoria, column = 1).value
            if data == None:
                is_data = False
        count_row_categoria -=1
        self.count_row_categoria = count_row_categoria
        
        is_data = True
        count_row_localidade_riachuelo = 1
        while is_data:
            count_row_localidade_riachuelo += 1
            data =  self.ws_categoria.cell (row = count_row_localidade_riachuelo, column = 1).value
            if data == None:
                is_data = False
        count_row_localidade_riachuelo -=1
        self.count_row_localidade_riachuelo = count_row_localidade_riachuelo
             
    def copia_dados(self):
        col_consolidado = 1
        col_fornecedor = 1
        linkado_fornecedor = 1
        linkado_consolidado = 1
        list_col_consolidado = []
        self.list_col_fornecedor = []
        list_linkado_fornecedor = []
        list_linkado_consolidado = []

        contador = 0
    
        while linkado_fornecedor <= self.count_row_linkado:
            nome_linkado_fornecedor =  self.ws_linkado.cell(row = linkado_fornecedor+1, column = 1).value
            list_linkado_fornecedor.append(nome_linkado_fornecedor)
            
            linkado_fornecedor+=1
        
        while linkado_consolidado <= self.count_row_linkado:
            nome_linkado_consolidado =  self.ws_linkado.cell(row = linkado_consolidado+1, column = 2).value
            list_linkado_consolidado.append(nome_linkado_consolidado)
            
            linkado_consolidado+=1
        
        while col_fornecedor <= self.count_col_fornecedor:
            nome_col1_fornecedor =  self.ws_fornecedor.cell(row = 1, column = col_fornecedor).value
            self.list_col_fornecedor.append(nome_col1_fornecedor)
            
            col_fornecedor+=1
            
        while col_consolidado <= self.count_col_consolidado:
            nome_col1_consolidado =  self.ws_consolidado.cell(row = 1, column = col_consolidado).value
            list_col_consolidado.append(nome_col1_consolidado)
            
            if nome_col1_consolidado in list_linkado_consolidado:
                indice_row_consolidado = list_col_consolidado.index(nome_col1_consolidado)
                                
                nome_fornecedor_indexado = list_linkado_fornecedor[indice_row_consolidado]
                
                for j in range(self.count_col_fornecedor):
                    cabeçalho_fornecedor = self.ws_fornecedor.cell(row = 1, column = j+1).value
                    if nome_fornecedor_indexado == cabeçalho_fornecedor:
                        for i in range(self.count_row_fornecedor):
                            contador+=1
                            valor_all_linha_cabeçalho = self.ws_fornecedor.cell(row = i+2, column = j+1).value
                            
                            self.ws_consolidado.cell(row = i+2, column = col_consolidado).value = valor_all_linha_cabeçalho            
                        
            col_consolidado+=1
        
        self.list_linkado_consolidado = list_linkado_consolidado
        self.list_col_consolidado = list_col_consolidado

    def exportar_twm_categoria_localidade(self):
        list_col_twm = []
        for col_twm in range (self.count_col_twm):
            list_col_twm.append(self.ws_twm.cell(row=1, column=col_twm+1).value)
        self.list_col_twm = list_col_twm
        
        try: #Dados do TWM
            #indice_conta_aglutinada = list_col_twm.index('Conta aglutinada') + 1
            indice_col_identificador = self.list_col_consolidado.index('Identificador') + 1
            #indice_num_conta = self.list_col_consolidado.index('Nº da Conta') + 1 
        except: # Dados do NEXinvoice
            #indice_conta_aglutinada = list_col_twm.index('CONTRATO') + 1
            indice_col_identificador = self.list_col_consolidado.index('FATURA') + 1
            #indice_num_conta = self.list_col_consolidado.index('CONTRATO') + 1 
            
        indice_desc_serviço = self.list_col_consolidado.index('Descrição Serviço') + 1  
        indice_endereço_riachuelo = self.list_col_consolidado.index('Endereço cliente') + 1
        indice_col_CNPJ_CPFL = self.list_col_consolidado.index('CNPJ Fornecedor') + 1
        #indice_col_vencimento = self.list_col_consolidado.index('Vencimento') + 1
        indice_nome_fornecedor = self.list_col_fornecedor.index('Nome_fornecedor') + 1
        
        self.Nome_do_fornecedor = self.ws_fornecedor.cell(row = 2, column = indice_nome_fornecedor).value
          
        j = 0
        
        while j < len(self.list_col_consolidado):
                              
            if self.list_col_consolidado[j] == 'Categoria':
                
                for i in range (self.count_row_fornecedor-1):
                    for c in range(self.count_row_categoria-1):
                        try:                           
                            descricao_servico = str(self.ws_consolidado.cell (row = i+2, column = indice_desc_serviço).value).upper() 
                            descricao_categoria = str(self.ws_categoria.cell (row =c+2, column = 1).value).upper() 
                            
                            descricao_servico = ''.join(filter(str.isalpha, descricao_servico))
                            descricao_categoria = ''.join(filter(str.isalpha, descricao_categoria))
                            
                            if descricao_servico == descricao_categoria:
                                self.ws_consolidado.cell (row = i+2, column = j+1).value = self.ws_categoria.cell (row =c+2, column = 2).value
                        except:
                            pass
#-------------------------------------------

            elif self.list_col_consolidado[j] == 'Subcategoria':
                for i in range (self.count_row_fornecedor-1):
                    for c in range(self.count_row_categoria-1):
                        if (str(self.ws_consolidado.cell (row = i+2, column = indice_desc_serviço).value)).upper() == (str(self.ws_categoria.cell (row =c+2, column = 1).value)).upper():
                           self.ws_consolidado.cell (row = i+2, column = j+1).value = self.ws_categoria.cell (row =c+2, column = 3).value
                indice_col_subcategoria = j + 1
            
            if self.Cliente == 'RIACHUELO':
                if self.list_col_consolidado[j] == 'Localidade':                    
                    for i in range (self.count_row_fornecedor-1):
                        for l in range (self.count_row_localidade_riachuelo):
                            #print(j, ':', self.list_col_consolidado[j],'----', i, '----', l)
                            if (str(self.ws_consolidado.cell (row = i+2, column = 1).value)).upper() == (str(self.ws_localidade_riachuelo.cell (row = l+2, column = 1).value)).upper():
                                self.ws_consolidado.cell (row = i+2, column = j+1).value = self.ws_localidade_riachuelo.cell (row = l+2, column = 4).value
                            
                            elif (str(self.ws_consolidado.cell (row = i+2, column = indice_endereço_riachuelo).value)[:15]).upper() in (str(self.ws_localidade_riachuelo.cell (row = l+2, column = 5).value)).upper():
                                self.ws_consolidado.cell (row = i+2, column = j+1).value = self.ws_localidade_riachuelo.cell (row = l+2, column = 4).value
                                
            if self.Cliente == 'FLEURY':
               if self.list_col_consolidado[j] == 'Localidade':
                   col_localidade_fleury = j
                   
               if self.list_col_consolidado[j] == 'Cód. Filial': 
                   for i in range (self.count_row_fornecedor-1):
                       self.ws_consolidado.cell (row = i+2, column = j+1).value = self.ws_consolidado.cell (row = i+2, column = col_localidade_fleury+1).value
                   
               if self.list_col_consolidado[j] == 'Mês':
                   for i in range (self.count_row_fornecedor-1):
                       self.ws_consolidado.cell (row = i+2, column = j+1).value = self.palavra_chave[3:6]  
            j += 1


        if self.Nome_do_fornecedor == 'CPFL':
            for i in range (self.count_row_fornecedor-1):
                self.ws_consolidado.cell (row = i+2, column = indice_col_CNPJ_CPFL).value = '04.172.213/0001-51'
        
        cliente = self.Cliente
        self.wb_consolidado.save(f'Consolidados\{cliente}/______consolidado_' + self.Cliente + '_' + self.Nome_do_fornecedor+'.xlsx')
        print(r'Save file >>>>>', '______consolidado_' + self.Cliente + '_' + self.Nome_do_fornecedor+'.xlsx')
        