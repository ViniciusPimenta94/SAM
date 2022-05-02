from pathlib import Path
import openpyxl, os, shutil, time
import pandas as pd
import numpy as np
        
class Validando_arquivos:
        def __init__(self, banco, cliente, inicio):
            self.banco = banco
            self.cliente = cliente
            self.inicio = inicio
            Validando_arquivos.validar_consumo(self)
        
        def num_float(self,num):

            if num == None:
                num = float(0)

            if type(num) == str:
                string = '!.#%'
                char_rep = {k: '' for k in string}
                num = num.translate(str.maketrans(char_rep))
                num = float(num.replace(',', '.'))
            return num
                
        def validar_consumo(self):
            myfile = Path(f'./Dashboard/{self.cliente}/Validação {self.cliente}.txt')
            myfile.touch(exist_ok=True)
            f = open(f'./Dashboard/{self.cliente}/Validação {self.cliente}.txt', 'w')
            
            try:
                # Carregando Excel
                print('\nVALIDANDO CONSUMO\n')
                wb_consolidado = openpyxl.load_workbook(f'./Dashboard/{self.cliente}/_____CONSOLIDADO_COMPLETO' + self.banco + '.xlsx')
                wb_consolidado.save('CONSOLIDADO_Consumo' + self.banco + '.xlsx')
                ws_consolidado = wb_consolidado['Sheet']
    
                a = 1
                b = 0
    
                # Encontrar índices da coluna
                while b < a:
                    try:
                        coluna = ws_consolidado.cell(row=1, column=a).value
                        if coluna == None:
                            a -= 1
                        if coluna == 'Categoria':
                            cat = a
                        if coluna == 'Subcategoria':
                            subcat = a
                        if coluna == 'Quantidade':
                            indice_quant = a
                        if coluna == 'Valor do Serviço':
                            valor = a
                    except:
                        pass
                    a += 1
                    b += 1
                
                is_data = True
                count_row_consolidado = 1
                while is_data:
                    count_row_consolidado += 1
                    data =  ws_consolidado.cell (row = count_row_consolidado, column = 1).value
                    if data == None:
                        is_data = False
                count_row_consolidado-=1
                
                c = count_row_consolidado
    
                i = 2
                valor_servico = 0
    
                # Filtrando Consumo
                while i <= c:
                    try:
                        ws_consolidado.cell (row = i, column=indice_quant).value = Validando_arquivos.num_float(self, ws_consolidado.cell (row = i, column=indice_quant).value) 
                        ws_consolidado.cell (row = i, column=valor).value = Validando_arquivos.num_float(self, ws_consolidado.cell (row = i, column=valor).value)
                    except:
                        pass
                    
                    categoria = ws_consolidado.cell(row=i, column=cat).value
                    subcategoria = ws_consolidado.cell(row=i, column=subcat).value
                    valor_servico += ws_consolidado.cell(row = i, column=valor).value
                    
                    if categoria == None:
                        categoria = '-'
                    if subcategoria == None:
                        subcategoria = '-'
                       
                    if categoria.lower() == 'consumo' and subcategoria.lower() == 'na ponta':
                        pass #print(categoria,'-', subcategoria)
                    elif categoria.lower() == 'consumo' and subcategoria.lower() == 'fora ponta':
                        pass #print(categoria,'-', subcategoria)
                    elif categoria.lower() == 'consumo' and subcategoria.lower() == 'te na ponta':
                        pass #print(categoria,'-', subcategoria)
                    elif categoria.lower() == 'consumo' and subcategoria.lower() == 'te fora ponta':
                        pass #print(categoria,'-', subcategoria)
                    elif categoria.lower() == 'consumo água':
                        pass
                    else:
                        ws_consolidado.cell(row=i, column=indice_quant).value = None
                        # print(categoria, ':', subcategoria, '=', None)
                        
                    i += 1
                
                wb_consolidado.save('CONSOLIDADO_Consumo' + self.banco + '.xlsx')
    
                # Criando tabelas dinâmicas
                tabela1 = pd.read_excel('CONSOLIDADO_Consumo' + self.banco + '.xlsx')
                tabela2 = pd.read_excel('./arquivo_twm/arquivo_TWM_' + self.cliente + '.xlsx')
                
                tabela_consolidado = pd.pivot_table(tabela1, index=['Identificador'], values=['Quantidade'], aggfunc=[np.sum])
                tabela_twm = pd.pivot_table(tabela2, index=['Identificador'], values=['Consumo da fatura'], aggfunc=[np.sum])
                tabela_twm_valor = pd.pivot_table(tabela2, index=['Identificador'], values=['Valor'], aggfunc=[np.sum])
                tabela_fornecedor = pd.pivot_table(tabela2, index=['Identificador'], values=['Fornecedor'], aggfunc=[np.sum])
    
                lista_id_consolidado = tabela_consolidado.index.tolist()
                lista_consumo_consolidado = tabela_consolidado.values.tolist()
                lista_id_twm = tabela_twm.index.tolist()
                lista_consumo_twm = tabela_twm.values.tolist()
                lista_id_fornecedor = tabela_fornecedor.index.tolist()
                lista_fornecedor = tabela_fornecedor.values.tolist()
                lista_twm_valor = tabela_twm_valor.values.tolist()
                
                for i in lista_id_twm:
                    if i in lista_id_consolidado:
                        pass
                    else:
                        lista_id_consolidado.append(i)
                        lista_consumo_consolidado.append(['Não encontrado'])
                        
                valor_twm = 0
                
                for i in lista_twm_valor:
                    valor_twm += float(i[0])
       
                dict_consolidado = dict(zip(lista_id_consolidado, lista_consumo_consolidado))
                dict_twm = dict(zip(lista_id_twm, lista_consumo_twm))
                dict_fornecedor = dict(zip(lista_id_fornecedor, lista_fornecedor))
                
                # print(dict_consolidado)
                # print(dict_twm)
                
                certo = 0
                check_consumo = []
    
                # Encontrando consumos errados
                for key in dict_twm.keys():
                    try:
                        consumo_twm = round(Validando_arquivos.num_float(self, dict_twm[key][0]),2)
                        consumo_consolidado = round(dict_consolidado[key][0], 2)
                        # print(consumo_twm, ' : ', consumo_consolidado)
                        
                        if consumo_twm == consumo_consolidado:
                            certo += 1
                            
                        if consumo_twm != consumo_consolidado:
                            check_consumo.append(dict_fornecedor[key][0], key, dict_twm[key][0], dict_consolidado[key][0])
                            # print(f'{key} - Fornecedor: {dict_fornecedor[key]} - TWM: {dict_twm[key]} - Consolidado: {dict_consolidado[key][0]}')
                        else:
                            pass #print(consumo_twm,' == ', consumo_consolidado)
                    except:
                        check_consumo.append([dict_fornecedor[key][0], key, dict_twm[key][0], dict_consolidado[key][0]])
                        # print(f'{key} - Fornecedor: {dict_fornecedor[key]} - Não encontrado')
                
                check_consumo.sort()
                print(f'{certo} faturas com consumo correto\n')
                print('Verificar as seguintes faturas: ')
                f.write(f'{certo} faturas com consumo correto\n')
                f.write('\nVerificar as seguintes faturas: \n')
                
                count = 0
                verificar = []
                for i in check_consumo:
                    if Validando_arquivos.num_float(self, i[2]) == 0 and i[3] != 'Não encontrado':
                        pass
                    elif i[3] == 'Não encontrado':
                        verificar.append(i)
                        count += 1
                    else:
                        count += 1
                        if i[3] == 'Não encontrado':
                            pass
                        else:
                            print(f'{i[0]} - {i[1]} - TWM [{i[2]}] - Consolidado [{i[3]}]')
                            f.writelines(f'\n{i[0]} - {i[1]} - TWM [{i[2]}] - Consolidado [{i[3]}]')
                print()
                f.write('\n')
                for i in verificar:
                    print(f'{i[0]} - {i[1]} - TWM [{i[2]}] - Consolidado [{i[3]}]')
                    f.writelines(f'\n{i[0]} - {i[1]} - TWM [{i[2]}] - Consolidado [{i[3]}]')
    
                print(f'\n{count} faturas para verificar')
                f.write(f'\n\n{count} faturas para verificar')
                f.write('\n')
            
            except:
                print('Não foi possível validar o consumo')
                print('Verificar se os nomes da colunas estão corretos. TWM [Identificador e Consumo da Fatura] - Consolidado [Identificador e Quantidade]')
                print('\nCaso seja um cliente NEX, não é possível fazer a validação de consumo, pois não há coluna de consumo na planilha do NEX')
            
#--------------------------------------------------------------------------------------------------------------------
            print('\nVALIDANDO CNPJ\n')
            wb_consolidado = openpyxl.load_workbook(f'./Dashboard/{self.cliente}/_____CONSOLIDADO_COMPLETO' + self.banco + '.xlsx')
            ws_consolidado = wb_consolidado['Sheet']
            
            tabela = pd.read_excel(f'./Dashboard/{self.cliente}/_____CONSOLIDADO_COMPLETO' + self.banco + '.xlsx')
            tabela_cnpjf = pd.pivot_table(tabela, index=['CNPJ Fornecedor'])
            tabela_cnpjc = pd.pivot_table(tabela, index=['CNPJ Cliente'])
            
            lista_cnpjf = tabela_cnpjf.index.tolist()
            lista_cnpjc = tabela_cnpjc.index.tolist()

            cnpjs = []
            cnpjs.append(list(lista_cnpjf))
            cnpjs.append(list(lista_cnpjc))

            for ind, value in enumerate(cnpjs):
                if ind == 0:
                    nome = 'FORNECEDOR'
                else:
                    nome = 'CLIENTE'
                    
                for i, v in enumerate(value):
                    if v == None:
                        pass
                    else:
                        cnpj = v
                    
                    lista = []
                    lista2 = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
                    soma1 = 0
                    soma2 = 0
                    
                    cnpjn = ''.join(filter(str.isnumeric, cnpj))
                    
                    for i, v in enumerate(cnpjn):
                        if i < 12:
                            lista.append(cnpjn[i])
                    
                    for i, v in enumerate(lista):
                        soma1 += int(lista[i])*(lista2[i])
                    
                    valor = 11 - (soma1 % 11)
                    
                    if valor > 9:
                        valor = '0'
                    else:
                        valor = str(valor)
                    
                    lista += valor
                    
                    x = [6]
                    x.extend(lista2)
                    
                    for i, v in enumerate(lista):
                        soma2 += int(lista[i])*(x[i])
                    
                    valor2 = 11 - (soma2 % 11)
                    
                    if valor2 > 9:
                        valor2 = '0'
                    else:
                        valor2 = str(valor2)
                    
                    lista += valor2
                    
                    sequencia = ''.join(cnpjn) == (str(lista[0])) * len(cnpjn)    # Elimina as sequências
                    
                    validar = ''.join(lista)
                    
                    if validar == cnpjn and not sequencia:
                        #pass
                        f.writelines(f'\n{cnpj} - CNPJ {nome} VÁLIDO')
                    else:
                        if cnpj == '' or cnpj == None:
                            pass
                        else:
                            print(f'{cnpj} - CNPJ {nome} INVÁLIDO')
                            f.writelines(f'\n{cnpj} - CNPJ {nome} INVÁLIDO')
            f.write('\n')                    
            os.remove('CONSOLIDADO_Consumo' + self.banco + '.xlsx')
            os.remove('_____TWM_'+self.banco+'.xlsx')
            
            try:
                shutil.rmtree('./PDF')
                shutil.rmtree('./Verticais')
            except OSError as e:
                print(f'\n{e}\n')
                
            try:                
                print(f'Total consolidado [{valor_servico:.2f}] : Total TWM [{valor_twm:.2f}]')
                f.writelines(f'\nTotal consolidado [{valor_servico:.2f}] : Total TWM [{valor_twm:.2f}]\n')
            except:
                tabela_twm_valor = pd.pivot_table(tabela2, index=['FATURA'], values=['VALOR'], aggfunc=[np.sum])
                lista_twm_valor = tabela_twm_valor.values.tolist()
                
                valor_twm = 0
                
                for i in lista_twm_valor:
                    valor_twm += float(i[0])
                    
                print(f'Total consolidado [{valor_servico:.2f}] : Total TWM [{valor_twm:.2f}]')
                f.writelines(f'\nTotal consolidado [{valor_servico:.2f}] : Total TWM [{valor_twm:.2f}]\n')

            print('-'*75)
            print('FIM DO PROCESSO')
            fim = time.time()
            total = (fim - self.inicio)
            tempo = time.strftime('%H:%M:%S', time.gmtime(total))
            f.write(f'\nO tempo de execução foi de {tempo}\n')
            f.write('-'*75)
            f.close()
