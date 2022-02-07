import sys
from PyQt5 import uic, QtWidgets
import openpyxl

class load_arquivos:
        def __init__(self):
            self.arquivo_fornecedor = []
            self.list_fornecedor = []
            self.wb_new = openpyxl.Workbook()
            self.ws_new = self.wb_new.active
            
            
        def ler_arquivo_fornecedor(self):
            
            fornecedor = QtWidgets.QFileDialog.getOpenFileName()[0]
            with open (fornecedor, 'r') as a:
                self.arquivo_fornecedor = a.name
            
            self.list_fornecedor.append(fornecedor)
            
            print('Load file fornecedor complete!')
            print('list_fornecedor:', self.list_fornecedor)

        def juntar_arquivos(self):
            k = 0
            ultima_linha_ws_new = 1
            
            while k<= len(self.list_fornecedor)-1:
                self.wb = openpyxl.load_workbook(self.list_fornecedor[k])
                self.ws = self.wb ['Worksheet']
                
                is_data = True
                count_row_ws = 1
                while is_data:
                    count_row_ws += 1
                    data =  self.ws.cell (row = count_row_ws, column = 1).value
                    if data == None:
                        is_data = False
                count_row_ws -=1
                
                is_data = True
                count_col_ws = 1
                while is_data:
                    count_col_ws += 1
                    data =  self.ws.cell (row = 1, column = count_col_ws).value
                    if data == None:
                        is_data = False
                count_col_ws -=1
                
                for j in range(count_col_ws):
                    for i in range(count_row_ws):
                        if k ==0:
                            i-=1
                        self.ws_new.cell(row=ultima_linha_ws_new+i+1, column= j+1).value = self.ws.cell(row=i+2,column=j+1).value
                ultima_linha_ws_new += count_row_ws-1
                k+=1

            self.wb_new.save('_____CONSOLIDADO_COMPLETO.xlsx')
            print('Complete!!')

app = QtWidgets.QApplication([])
tela = uic.loadUi("Interface_SAM.ui")

tela.show()

load_active = load_arquivos()
tela.pushButton.clicked.connect(load_active.ler_arquivo_fornecedor)
tela.pushButton_4.clicked.connect(load_active.juntar_arquivos)

sys.exit(app.exec_())