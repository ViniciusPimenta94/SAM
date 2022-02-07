import openpyxl

class consolidando_arquivos:
    def __init__(self, wb_fornecedor, Cliente, palavra_chave):
        print('Executando SAM_3...')
        self.Cliente = Cliente
        self.palavra_chave = palavra_chave
                
        self.wb_fornecedor = wb_fornecedor
        self.ws_fornecedor = self.wb_fornecedor['Worksheet2']
        
        self.wb_twm = openpyxl.load_workbook('arquivo_TWM_'+self.Cliente+'.xlsx')
        self.ws_twm = self.wb_twm['Planilha1']
        
        self.wb_categoria = openpyxl.load_workbook('Categorias_Subcategoria.xlsx')
        self.ws_categoria = self.wb_categoria['Categorias']
        
        self.wb_localidade_riachuelo = openpyxl.load_workbook('Localidades_RIACHUELO.xlsx')
        self.ws_localidade_riachuelo = self.wb_localidade_riachuelo['Localidades']
                
    def ler_arquivo_consolidado(self):
        if self.Cliente == 'FLEURY':
            self.arquivo_consolidado = 'arquivo_consolidado_FLEURY.xlsx'
            print('Load file consolidado complete!')
        
        elif self.Cliente == 'LOCALIZA':
            self.arquivo_consolidado = 'arquivo_consolidado_LOCALIZA.xlsx'
            print('Load file consolidado complete!')
            
        elif self.Cliente == 'PUC':
            self.arquivo_consolidado = 'arquivo_consolidado_PUC.xlsx'
            print('Load file consolidado complete!')
            
        elif self.Cliente == 'RIACHUELO':
            self.arquivo_consolidado = 'arquivo_consolidado_RIACHUELO.xlsx'
            print('Load file consolidado complete!')
        
        elif self.Cliente == 'DAKI':
            self.arquivo_consolidado = 'arquivo_consolidado_DAKI.xlsx'
            print('Load file consolidado complete!')
            
        else: 
            print('Arquivo consolidado não encontrado!!!')
            self.arquivo_consolidado = None
   
    def ler_arquivo_linkado(self):
        if self.Cliente == 'FLEURY':
            self.arquivo_linkado = 'arquivo_linkado_FLEURY.xlsx'
            print('Load file linkado complete!')
        
        elif self.Cliente == 'LOCALIZA':
            self.arquivo_linkado = 'arquivo_linkado_LOCALIZA.xlsx'
            print('Load file linkado complete!')
            
        elif self.Cliente == 'PUC':
            self.arquivo_linkado = 'arquivo_linkado_PUC.xlsx'
            print('Load file linkado complete!')
            
        elif self.Cliente == 'RIACHUELO':
            self.arquivo_linkado = 'arquivo_linkado_RIACHUELO.xlsx'
            print('Load file linkado complete!')
            
        elif self.Cliente == 'DAKI':
            self.arquivo_linkado = 'arquivo_linkado_DAKI.xlsx'
            print('Load file linkado complete!')
        
        else: 
            print('Arquivo Linkado não encontrado!!!')
            self.arquivo_linkado = None
 
    def load_wb_Consolidado(self):
        wb_consolidado = openpyxl.load_workbook(self.arquivo_consolidado)
        ws_consolidado = wb_consolidado['Worksheet']
        
        self.wb_consolidado = wb_consolidado
        self.ws_consolidado = ws_consolidado

    def load_wb_linkado(self):
        wb_linkado = openpyxl.load_workbook(self.arquivo_linkado)
        ws_linkado = wb_linkado['Worksheet']

        self.wb_linkado = wb_linkado
        self.ws_linkado = ws_linkado

    def count_colunas(self):
        is_data = True
        count_col_consolidado = 1
        while is_data:
            count_col_consolidado += 1
            data =  self.ws_consolidado.cell (row = 1, column = count_col_consolidado).value
            if data == None:
                is_data = False
        count_col_consolidado-=1
        self.count_col_consolidado = count_col_consolidado
        
        is_data = True
        count_col_fornecedor = 1
        while is_data:
            count_col_fornecedor += 1
            data =  self.ws_fornecedor.cell (row = 1, column = count_col_fornecedor).value
            if data == None:
                is_data = False
        count_col_fornecedor-=1
        self.count_col_fornecedor = count_col_fornecedor
        
        is_data = True
        count_col_twm = 1
        while is_data:
            count_col_twm += 1
            data =  self.ws_twm.cell (row = 1, column = count_col_twm).value
            if data == None:
                is_data = False
        count_col_twm-=1
        self.count_col_twm = count_col_twm                

    def count_linhas(self):
        is_data = True
        count_row_fornecedor = 1
        while is_data:
            count_row_fornecedor += 1
            data =  self.ws_fornecedor.cell (row = count_row_fornecedor, column = 1).value
            if data == None:
                is_data = False
        count_row_fornecedor -=1
        self.count_row_fornecedor = count_row_fornecedor
        
        is_data = True
        count_row_linkado = 1
        while is_data:
            count_row_linkado += 1
            data =  self.ws_linkado.cell (row = count_row_linkado, column = 1).value
            if data == None:
                is_data = False
        count_row_linkado -=1
        self.count_row_linkado = count_row_linkado
        
        is_data = True
        count_row_twm = 1
        while is_data:
            count_row_twm += 1
            data =  self.ws_twm.cell (row = count_row_twm, column = 1).value
            if data == None:
                is_data = False
        count_row_twm -=1
        self.count_row_twm = count_row_twm
        
        is_data = True
        count_row_categoria = 1
        while is_data:
            count_row_categoria += 1
            data =  self.ws_categoria.cell (row = count_row_categoria, column = 1).value
            if data == None:
                is_data = False
        count_row_categoria -=1
        self.count_row_categoria = count_row_categoria
        
        is_data = True
        count_row_localidade_riachuelo = 1
        while is_data:
            count_row_localidade_riachuelo += 1
            data =  self.ws_categoria.cell (row = count_row_localidade_riachuelo, column = 1).value
            if data == None:
                is_data = False
        count_row_localidade_riachuelo -=1
        self.count_row_localidade_riachuelo = count_row_localidade_riachuelo
             
    def copia_dados(self):
        col_consolidado = 1
        col_fornecedor = 1
        linkado_fornecedor = 1
        linkado_consolidado = 1
        list_col_consolidado = []
        self.list_col_fornecedor = []
        list_linkado_fornecedor = []
        list_linkado_consolidado = []

        contador = 0
    
        while linkado_fornecedor <= self.count_row_linkado:
            nome_linkado_fornecedor =  self.ws_linkado.cell(row = linkado_fornecedor+1, column = 1).value
            list_linkado_fornecedor.append(nome_linkado_fornecedor)
            
            linkado_fornecedor+=1
        
        while linkado_consolidado <= self.count_row_linkado:
            nome_linkado_consolidado =  self.ws_linkado.cell(row = linkado_consolidado+1, column = 2).value
            list_linkado_consolidado.append(nome_linkado_consolidado)
            
            linkado_consolidado+=1
        
        while col_fornecedor <= self.count_col_fornecedor:
            nome_col1_fornecedor =  self.ws_fornecedor.cell(row = 1, column = col_fornecedor).value
            self.list_col_fornecedor.append(nome_col1_fornecedor)
            
            col_fornecedor+=1
            
        while col_consolidado <= self.count_col_consolidado:
            nome_col1_consolidado =  self.ws_consolidado.cell(row = 1, column = col_consolidado).value
            list_col_consolidado.append(nome_col1_consolidado)
            
            if nome_col1_consolidado in list_linkado_consolidado:
                indice_row_consolidado = list_col_consolidado.index(nome_col1_consolidado)
                                
                nome_fornecedor_indexado = list_linkado_fornecedor[indice_row_consolidado]
                
                for j in range(self.count_col_fornecedor):
                    cabeçalho_fornecedor = self.ws_fornecedor.cell(row = 1, column = j+1).value
                    if nome_fornecedor_indexado == cabeçalho_fornecedor:
                        for i in range(self.count_row_fornecedor):
                            contador+=1
                            valor_all_linha_cabeçalho = self.ws_fornecedor.cell(row = i+2, column = j+1).value
                            
                            self.ws_consolidado.cell(row = i+2, column = col_consolidado).value = valor_all_linha_cabeçalho            
                        
            col_consolidado+=1
        
        self.list_linkado_consolidado = list_linkado_consolidado
        self.list_col_consolidado = list_col_consolidado

    def exportar_twm_categoria_localidade(self):
        list_col_twm = []
        for col_twm in range (self.count_col_twm):
            list_col_twm.append(self.ws_twm.cell(row=1, column=col_twm+1).value)
        self.list_col_twm = list_col_twm
        
        indice_conta_aglutinada = list_col_twm.index('Conta aglutinada') + 1
        indice_desc_serviço = self.list_col_consolidado.index('Descrição Serviço') + 1  
        indice_num_conta = self.list_col_consolidado.index('Nº da Conta') + 1 
        indice_endereço_riachuelo = self.list_col_consolidado.index('Endereço cliente') + 1
        indice_col_CNPJ_CPFL = self.list_col_consolidado.index('CNPJ Fornecedor') + 1
        indice_col_vencimento = self.list_col_consolidado.index('Vencimento') + 1
        
        indice_nome_fornecedor = self.list_col_fornecedor.index('Nome_fornecedor') + 1
        self.Nome_do_fornecedor = self.ws_fornecedor.cell(row = 2, column = indice_nome_fornecedor).value

####################################################################################        
        
        list_identificador_teste = []
        for i in range (self.count_row_fornecedor-1):
            conta_aglutinada = str(self.ws_consolidado.cell(row = i+2, column = indice_num_conta).value)
            vencimento = str(self.ws_consolidado.cell(row = i+2, column = indice_col_vencimento).value)
            
            ano = vencimento [6:]
            mes = vencimento [3:5]
            dia = vencimento [:2]
            
            identificador_teste = conta_aglutinada + '_'+ ano + mes + dia
            list_identificador_teste.append(identificador_teste)
            
####################################################################################
        
        j = 0
        while j < len(self.list_col_consolidado):
            print(j, ':', self.list_col_consolidado[j])

            if self.list_col_consolidado[j] in list_col_twm:
                indice_twm = list_col_twm.index(self.list_col_consolidado[j]) + 1                
                
                for i in range (self.count_row_fornecedor-1):
                      for t in range (self.count_row_twm-1):          
                            if self.ws_consolidado.cell (row = i+2, column = indice_num_conta).value == self.ws_twm.cell (row =t+2, column = indice_conta_aglutinada).value:
                              self.ws_consolidado.cell (row = i+2, column = j+1).value = self.ws_twm.cell (row =t+2, column = indice_twm).value


            if self.list_col_consolidado[j] == 'Categoria':
                for i in range (self.count_row_fornecedor-1):
                    for c in range(self.count_row_categoria-1):
                        #print(j, ':', self.list_col_consolidado[j],'----', i, '----', c)
                        if (str(self.ws_consolidado.cell (row = i+2, column = indice_desc_serviço).value)).upper() == (str(self.ws_categoria.cell (row =c+2, column = 1).value)).upper():
                            self.ws_consolidado.cell (row = i+2, column = j+1).value = self.ws_categoria.cell (row =c+2, column = 2).value
                        elif (str(self.ws_categoria.cell (row =c+2, column = 4).value)).upper() in (str(self.ws_consolidado.cell (row = i+2, column = indice_desc_serviço).value)).upper():
                              self.ws_consolidado.cell (row = i+2, column = j+1).value = self.ws_categoria.cell (row =c+2, column = 2).value

            elif self.list_col_consolidado[j] == 'Subcategoria':
                for i in range (self.count_row_fornecedor-1):
                    for c in range(self.count_row_categoria-1):
                        #print(j, ':', self.list_col_consolidado[j],'----', i, '----', c)
                        if (str(self.ws_consolidado.cell (row = i+2, column = indice_desc_serviço).value)).upper() == (str(self.ws_categoria.cell (row =c+2, column = 1).value)).upper():
                           self.ws_consolidado.cell (row = i+2, column = j+1).value = self.ws_categoria.cell (row =c+2, column = 3).value
                        elif (str(self.ws_categoria.cell (row =c+2, column = 5).value)).upper() in (str(self.ws_consolidado.cell (row = i+2, column = indice_desc_serviço).value)).upper():
                             self.ws_consolidado.cell (row = i+2, column = j+1).value = self.ws_categoria.cell (row =c+2, column = 3).value
            
            if self.Cliente == 'RIACHUELO':
                if self.list_col_consolidado[j] == 'Localidade':                    
                    for i in range (self.count_row_fornecedor-1):
                        for l in range (self.count_row_localidade_riachuelo):
                            #print(j, ':', self.list_col_consolidado[j],'----', i, '----', l)
                            if (str(self.ws_consolidado.cell (row = i+2, column = 1).value)).upper() == (str(self.ws_localidade_riachuelo.cell (row = l+2, column = 1).value)).upper():
                                self.ws_consolidado.cell (row = i+2, column = j+1).value = self.ws_localidade_riachuelo.cell (row = l+2, column = 4).value
                            
                            elif (str(self.ws_consolidado.cell (row = i+2, column = indice_endereço_riachuelo).value)[:15]).upper() in (str(self.ws_localidade_riachuelo.cell (row = l+2, column = 5).value)).upper():
                                self.ws_consolidado.cell (row = i+2, column = j+1).value = self.ws_localidade_riachuelo.cell (row = l+2, column = 4).value
                                
            if self.Cliente == 'FLEURY':
               if self.list_col_consolidado[j] == 'Localidade':
                   col_localidade_fleury = j
                   
               if self.list_col_consolidado[j] == 'Cód. Filial': 
                   for i in range (self.count_row_fornecedor-1):
                       self.ws_consolidado.cell (row = i+2, column = j+1).value = self.ws_consolidado.cell (row = i+2, column = col_localidade_fleury+1).value
                   
               if self.list_col_consolidado[j] == 'Mês':
                   for i in range (self.count_row_fornecedor-1):
                       self.ws_consolidado.cell (row = i+2, column = j+1).value = self.palavra_chave[3:6]  
            j += 1

        if self.Nome_do_fornecedor == 'CPFL':
            for i in range (self.count_row_fornecedor-1):
                self.ws_consolidado.cell (row = i+2, column = indice_col_CNPJ_CPFL).value = '04.172.213/0001-51'
        
        self.wb_consolidado.save('______consolidado_' + self.Cliente + '_' + self.Nome_do_fornecedor+'.xlsx')
        print('Save file >>>>>', '______consolidado_' + self.Cliente + '_' + self.Nome_do_fornecedor+'.xlsx')
        
        
        #try:
        #   print('Valindando consumo...')
        #consolidando_arquivos.validar_consumo(self)
        #except:
        #    print('Valor consumo não pode ser avaliado!')

    def validar_consumo(self):
        j = 0
        while j <= (self.count_col_consolidado-1):
            if 'Consumo Faturado' == self.list_col_consolidado[j] or 'Consumo' == self.list_col_consolidado[j] or 'Quantidade' == self.list_col_consolidado[j]:
                indice_consumo_consolidado = j + 1
            if 'Identificador' in self.list_col_consolidado[j]: 
                indice_identificador_consolidado = j + 1                
            if 'Categoria' in self.list_col_consolidado[j]:
                indice_categoria_consolidado = j + 1   
            if 'Subcategoria' in self.list_col_consolidado[j]:
                    indice_subcategoria_consolidado = j + 1  
            if 'Valor total' in self.list_col_consolidado[j]:
                self.indice_valor_total = j + 1
                
            j+=1

        self.indice_consumo_consolidado = indice_consumo_consolidado
        
        j = 0
        while j<= (self.count_col_twm-1):
            if 'Consumo da fatura' in self.list_col_twm[j]: 
                indice_consumo_twm = j + 1
            if 'Identificador' in self.list_col_twm[j]: 
                indice_identificador_twm = j + 1   
            j+=1             
            
        i = 2
        l = 2
        valor_consumo = []
        id1 = []
        id2 = []
        
        while i <= (self.count_row_fornecedor):
            id_consolidado = self.ws_consolidado.cell(row=i, column=indice_identificador_consolidado).value
            
            categoria_consolidado = self.ws_consolidado.cell(row=i, column=indice_categoria_consolidado).value
            sub_categoria_consolidado = self.ws_consolidado.cell(row=i, column=indice_subcategoria_consolidado).value      
                
            if categoria_consolidado == None:
                categoria_consolidado = []
            if sub_categoria_consolidado == None: 
                 sub_categoria_consolidado = []
            
            id1 = id_consolidado
            
            if 'Consumo' in categoria_consolidado and 'Na ponta' in sub_categoria_consolidado or 'Consumo' in categoria_consolidado and 'Fora ponta' in sub_categoria_consolidado or 'Consumo' in categoria_consolidado and 'TE fora ponta' in sub_categoria_consolidado or 'Consumo' in categoria_consolidado and 'TE na ponta' in sub_categoria_consolidado:
                if (self.ws_consolidado.cell(row=i, column=indice_consumo_consolidado).value) != None:
                    
                    valor_consumo1 = self.ws_consolidado.cell(row=i, column=indice_consumo_consolidado).value
                    
                    if valor_consumo1 == None:
                        valor_consumo1 = float(0)
                        
                    if valor_consumo1[len(valor_consumo1)-2:len(valor_consumo1)-1] == '-':
                        print('simmmmmmm')
                        valor_consumo1 = valor_consumo1[:len(valor_consumo1)-1]
                        valor_consumo1 = '-'+valor_consumo1
                        valor_consumo1 = float(valor_consumo1)
                    
                    if valor_consumo1!= None and type(valor_consumo1) == str:
                        string2 = '!.#'
                        char_rep={k: '' for k in string2}
                        valor_consumo1=valor_consumo1.translate(str.maketrans(char_rep))
                        valor_consumo1 = valor_consumo1.replace (',','.') 
                    valor_consumo1 = float(valor_consumo1)
                    
                    valor_consumo.append(float(valor_consumo1)) 
                     

                    if id1 != id2:    
                        valor_consumo_consolidado = float(sum(valor_consumo))   
                        while l <= (self.count_row_twm):   
                            id_twm = self.ws_twm.cell(row=l, column=indice_identificador_twm).value
                            valor_consumo_twm = float(self.ws_twm.cell(row=l, column=indice_consumo_twm).value)
                            
                            if valor_consumo_twm == None:
                                valor_consumo_twm = float(0)
                            
                            if valor_consumo_twm!= None and type(valor_consumo_twm) == str:
                                string2 = '!.#'
                                char_rep={k: '' for k in string2}
                                valor_consumo_twm=valor_consumo_twm.translate(str.maketrans(char_rep))
                                valor_consumo_twm = valor_consumo_twm.replace (',','.') 
                            valor_consumo_twm = float(valor_consumo_twm)
                            
                            if id_consolidado == id_twm:
                                if valor_consumo_consolidado == valor_consumo_twm:
                                    print('valor consumo correto!', 'id:', id_consolidado)
                                    print('valor_consumo_consolidado', valor_consumo_consolidado)
                                    print('valor_consumo_twm', valor_consumo_twm)
                                else:
                                    print('valor consumo errado!', 'id:', id_consolidado)
                                    print('valor_consumo_consolidado', valor_consumo_consolidado)
                                    print('valor_consumo_twm', valor_consumo_twm)
                                    
                                    #consolidando_arquivos.ic_consumo(self, i)
                            l+=1
                        valor_consumo = []   
                        
                    elif i == (self.count_row_fornecedor):
                        if valor_consumo == None:
                            valor_consumo = float(0)
                        print('ultimo')
                        valor_consumo_consolidado = float(sum(valor_consumo))
                        print('valor_consumo_consolidado', valor_consumo_consolidado)
                        print('id_consolidado', id_consolidado)
                        l = 2
                        while l <= (self.count_row_twm):   
                            id_twm = self.ws_twm.cell(row=l, column=indice_identificador_twm).value
                            valor_consumo_twm = float(self.ws_twm.cell(row=l, column=indice_consumo_twm).value)
                            
                            if valor_consumo_twm == None:
                                valor_consumo_twm = float(0)
                            
                            if id_consolidado == id_twm:
                                if valor_consumo_consolidado == valor_consumo_twm:
                                    print('valor consumo correto!', 'id:', id_consolidado)
                                    print('valor_consumo_consolidado', valor_consumo_consolidado)
                                    print('valor_consumo_twm', valor_consumo_twm)
                                else:
                                    print('valor consumo errado!', 'id:', id_consolidado)
                                    print('valor_consumo_consolidado', valor_consumo_consolidado)
                                    print('valor_consumo_twm', valor_consumo_twm)
                                    
                                   # consolidando_arquivos.ic_consumo(self, i)
                            l+=1
                    id2 = id1
                else: 
                    None
            i+=1
            
    def ic_consumo(self, i):
        ic_superior = 2.5
        ic_inferior = 0.2
        
        valor = self.ws_consolidado.cell(row=i, column=self.indice_valor_total).value
        if valor!= None and type(valor) == str:
            string2 = '!.#'
            char_rep={k: '' for k in string2}
            valor=valor.translate(str.maketrans(char_rep))
            valor = valor.replace (',','.') 
        valor = float(valor)
        
        consumo = self.ws_consolidado.cell(row=i, column=self.indice_consumo_consolidado).value
        
        if consumo!= None and type(consumo) == str:
            string2 = '!.#'
            char_rep={k: '' for k in string2}
            consumo=consumo.translate(str.maketrans(char_rep))
            consumo = consumo.replace (',','.')
        
        if consumo != None:
            consumo = float(consumo)
        else:
            consumo = float(0)
        
        print('valor', valor, type(valor))
        print('consumo', consumo, type(consumo))
        
        razao_valor_consumo = float(valor/consumo)
        print('razao_valor_consumo', razao_valor_consumo)
        
        if ic_inferior <= razao_valor_consumo <= ic_superior:
            print('razao_valor_consumo',razao_valor_consumo, ' aceitável!!!')
        else:
            print('razao_valor_consumo', razao_valor_consumo, 'tem que analisar!!')
            

            
        