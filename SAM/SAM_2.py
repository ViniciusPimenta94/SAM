import openpyxl
from SAM_3 import consolidando_arquivos

class verificação():
    def __init__(self, arquivo_fornecedor, palavra_chave, cliente):
        self.arquivo_fornecedor = arquivo_fornecedor
        self.palavra_chave = palavra_chave
        self.cliente = cliente
        print('palavra_chave', palavra_chave)
    
    def criar_new_sheet(self):
        self.wb_fornecedor = openpyxl.load_workbook(self.arquivo_fornecedor)
        self.ws_fornecedor = self.wb_fornecedor ['Worksheet']
        self.ws_new_fornecedor = self.wb_fornecedor.create_sheet (title = "Worksheet2")
        
    def count_ws(self):
        is_data = True
        count_row_ws = 1
        while is_data:
            count_row_ws += 1
            data = self.ws_fornecedor.cell(row = count_row_ws, column = 1).value
            if data == None:
                is_data = False
        self.count_row_ws = count_row_ws

        is_data = True
        count_col_ws = 1
        while is_data:
            count_col_ws += 1
            data = self.ws_fornecedor.cell(row = 1, column = count_col_ws).value
            if data == None:
                is_data = False
        self.count_col_ws = count_col_ws

    def File_name(self):
        today_data = []
        i = 2
        r = 2
        mc = self.ws_fornecedor.max_column
        
        for j in range (1, mc+1):
            c = (self.ws_fornecedor.cell(row = 1, column = j)).value
            self.ws_new_fornecedor.cell(row = 1, column = j).value = c
            
        while i<=self.count_row_ws-1:
            row_data = {}
            row_data = self.ws_fornecedor.cell(row=i, column=1).value
            
            if self.palavra_chave in row_data:  
                for j in range (1, mc + 1): 
                      c = (self.ws_fornecedor.cell(row = i, column = j)).value                                      
                      self.ws_new_fornecedor.cell(row = r, column = j).value = c
                r+=1
                i+=1
            else:
                i+=1
                row_data = None
                #self.count_row_ws -= 1
            today_data.append(row_data)
            self.today_data = today_data
            
    def count_new_ws(self):
        is_data = True
        count_row_new_ws = 1
        while is_data:
            count_row_new_ws += 1
            data = self.ws_new_fornecedor.cell(row = count_row_new_ws, column = 1).value
            if data == None:
                is_data = False
        self.count_row_new_ws = count_row_new_ws
        
        is_data = True
        count_col_new_ws = 1
        while is_data:
            count_col_new_ws += 1
            data = self.ws_new_fornecedor.cell(row = 1, column = count_col_new_ws).value
            if data == None:
                is_data = False
        self.count_col_new_ws = count_col_new_ws      
          
          
    def tabela_dinâmica_new_sheet(self):
        list_identificador = []
        list_valor_total = []
        list_valor_auditoria = []
        list_layout = []
        
        col_fornecedor = []   
        t=1
        while t <= self.count_col_ws:
            col_data = self.ws_fornecedor.cell(row=1, column=t).value
            col_fornecedor.append(col_data)
            t+=1
        
        if 'Filename' in col_fornecedor:
            self.indice_filename = col_fornecedor.index('Filename') + 1
        
        if 'Dc_indentificador_layout' in col_fornecedor:
            self.indice_layout_col = col_fornecedor.index('Dc_indentificador_layout') + 1
        
        if 'Dc_identificador_layout' in col_fornecedor:
            self.indice_layout_col = col_fornecedor.index('Dc_identificador_layout') + 1
        
        for i in range(2, self.count_row_new_ws):
            identificador = self.ws_new_fornecedor.cell(row = i, column = self.indice_filename).value
            layout = self.ws_new_fornecedor.cell(row = i, column = self.indice_layout_col).value
            
            if identificador not in list_identificador:
                list_identificador.append(identificador)
                list_layout.append(layout)
        
        for i in range(len(list_identificador)):
            list_valor_auditoria.append(float(0)) 

        if 'Vl_total' in col_fornecedor:
            self.indice_vl_total_col = col_fornecedor.index('Vl_total') + 1
            
            for i in range(2, self.count_row_new_ws):
                Vl_string = self.ws_new_fornecedor.cell(row = i, column = self.indice_vl_total_col).value
                
                if Vl_string == None:
                    Vl_string = float(0)
               
                if type(Vl_string) == str:
                    string2 = '!.#%'
                    char_rep = {k: '' for k in string2}
                    Vl_string = Vl_string.translate(str.maketrans(char_rep))
                    Vl_total = float(Vl_string.replace (',','.'))
                
                else:
                    Vl_total = float(Vl_string)
                
                id1 = self.ws_new_fornecedor.cell(row=i-1, column=1).value
                id2 = self.ws_new_fornecedor.cell(row=i, column=1).value
                
                if id1 != id2:
                    
                    Vl_total = "{:.2f}".format(Vl_total)
                    
                    list_valor_total.append(float(Vl_total))
                    id2 = id1
                                
            self.soma_valor_total = float(sum(list_valor_total))  
        
        
        if 'Valores_faturados_auditoria Valor' in col_fornecedor: 
            self.indice_vl_faturado_col = col_fornecedor.index('Valores_faturados_auditoria Valor') + 1
            indice_auditoria = 0
            for i in range(2, self.count_row_new_ws):
                Faturado_string = self.ws_new_fornecedor.cell(row = i, column = self.indice_vl_faturado_col).value    
                
                if Faturado_string == None:
                    Faturado_string = float(0)
               
                if type(Faturado_string) == str:
                    string2 = '!.#%'
                    char_rep = {k: '' for k in string2}
                    Faturado_string = Faturado_string.translate(str.maketrans(char_rep))
                    vl_auditoria = float(Faturado_string.replace (',','.'))

                else:
                    vl_auditoria = float(Faturado_string)
                    
                id1 = self.ws_new_fornecedor.cell(row=i-1, column=1).value  
                id2 = self.ws_new_fornecedor.cell(row=i, column=1).value
                
                if i==2:
                    list_valor_auditoria[0] = vl_auditoria
                    
                elif id1 == id2:
                    if i == self.count_row_new_ws:
                        break
                    list_valor_auditoria[indice_auditoria] += vl_auditoria
                    list_valor_auditoria[indice_auditoria] = "{:.2f}".format(list_valor_auditoria[indice_auditoria])
                    list_valor_auditoria[indice_auditoria] = float(list_valor_auditoria[indice_auditoria])
                    
                else:  
                    indice_auditoria += 1
                    
                    if indice_auditoria == len(list_valor_auditoria):
                        indice_auditoria = len(list_valor_auditoria)
                    
                    list_valor_auditoria[indice_auditoria] += float(vl_auditoria)
                    list_valor_auditoria[indice_auditoria] = "{:.2f}".format(list_valor_auditoria[indice_auditoria])
                    list_valor_auditoria[indice_auditoria] = float(list_valor_auditoria[indice_auditoria])
                            
        print('-----id-------')
        print('-----',len(list_identificador),'-------')
        print(list_identificador)
        print('-----vlt------')
        print('-----',len(list_valor_total),'-------')
        print(list_valor_total)
        print('-----aud------')
        print('-----',len(list_valor_auditoria),'-------')
        print(list_valor_auditoria)
        print('-----aud------')
        print('-----',len(list_layout),'-------')
        print(list_layout)
        print('--------------')
        
        #disc_listas = dict.list_identificador
        
        
        
        

        dic_vltotal = dict(zip(list_identificador, list_valor_total))
        dic_auditoria = dict(zip(list_identificador, list_valor_auditoria))
        
        for i in range(len(list_identificador)):
            if dic_vltotal[list_identificador[i]] == dic_auditoria[list_identificador[i]]:    
                None
            else:       
                print('errado', list_identificador[i])
                print('layout errado = ', list_layout[i])

        self.soma_faturados_data = float(sum(list_valor_auditoria))
        self.soma_total_data = float(sum(list_valor_total))
        
    def comparar_valores(self):
        i=2   
        while i<=self.count_row_new_ws-1:
            c = (self.ws_new_fornecedor.cell(row = i, column = 1)).value                                      
            self.ws_new_fornecedor.cell(row = i, column = 1).value = c [:-12]
            i+=1
        
        std = self.wb_fornecedor.get_sheet_by_name('Worksheet')
        self.wb_fornecedor.remove_sheet(std)
        
        
        if self.soma_total_data == self.soma_faturados_data:
                self.analise = "Valor correto!!!"
                print('\n -------------- \n soma_total_data', self.soma_total_data)
                print('soma_total_faturado', self.soma_faturados_data)
                print(self.analise)
                
                chama_SAM3 = consolidando_arquivos (self.wb_fornecedor, self.cliente, self.palavra_chave)
                chama_SAM3.ler_arquivo_consolidado()
                chama_SAM3.ler_arquivo_linkado()
                chama_SAM3.load_wb_Consolidado()
                chama_SAM3.load_wb_linkado()
                chama_SAM3.count_colunas()
                chama_SAM3.count_linhas()
                chama_SAM3.copia_dados()
                chama_SAM3.exportar_twm_categoria_localidade()
                #self.wb_fornecedor.save(self.arquivo_fornecedor)
                #print('Save file fornecedor!!')
        else:
                self.analise = "Precisa verificar!!!"
                print(self.analise)
                print('soma_total_data', self.soma_total_data)
                print('soma_total_faturado', self.soma_faturados_data)
                #self.wb_fornecedor.save(self.arquivo_fornecedor)
                print('NO Save file fornecedor!!')
              
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
          
          
          
          
          
            